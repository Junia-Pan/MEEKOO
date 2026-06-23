# -*- coding: utf-8 -*-
"""
生成 MEEKOO-TMS 业务验收 Excel 模板（总览 + 清单两个 Sheet）。

用法：
  python admin/docs/_generate_uat_excel.py

输出：
  admin/docs/MEEKOO-TMS业务验收检查表.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent / "MEEKOO-TMS业务验收检查表.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
META_FONT = Font(size=10, color="444444")
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

RESULT_OPTIONS = '"通过,不通过,阻塞,待验证,N/A"'
LEVEL_OPTIONS = '"P0,P1,P2"'
STATUS_OPTIONS = '"待修复,修复中,待复验,已关闭"'

CHECK_HEADERS = [
    "场景编号", "验收分类", "验收阶段", "模块", "验收项",
    "预期结果/操作要点", "负责部门", "验收人", "验收日期", "验收结果", "备注/缺陷编号",
]

ALL_CHECK_ITEMS = [
    # 基础配置
    ("CFG-01", "基础配置", "阶段1-基础配置", "仓库设置", "仓库列表：新建/编辑/启用停用", "仓库信息保存正确，列表可查询", "仓库主管"),
    ("CFG-02", "基础配置", "阶段1-基础配置", "仓库设置", "仓库分区：分区与仓库关联", "分区创建后可被库位引用", "仓库主管"),
    ("CFG-03", "基础配置", "阶段1-基础配置", "仓库设置", "库位管理：库位创建与二维码", "库位二维码可被 PDA 扫描", "仓库主管"),
    ("CFG-04", "基础配置", "阶段1-基础配置", "基础数据", "客户管理：新建/编辑客户", "客户信息完整，可被计划引用", "运营"),
    ("CFG-05", "基础配置", "阶段1-基础配置", "基础数据", "供应商管理：提货/派送供应商", "供应商可被外仓/提拆派引用", "运营"),
    ("CFG-06", "基础配置", "阶段1-基础配置", "基础数据", "派送仓库/FBA仓配置", "仓码与分区配置正确", "运营"),
    ("CFG-07", "基础配置", "阶段1-基础配置", "基础数据", "船司/邮编/费用项/银行账号", "基础数据可被业务模块引用", "运营/财务"),
    ("CFG-08", "基础配置", "阶段1-基础配置", "系统配置", "组织架构/角色/用户/字典", "各岗位登录后菜单与权限正确", "管理员"),
    ("CFG-09", "基础配置", "阶段1-基础配置", "应收报价", "提拆派报价：新建五类价目", "价目 Tab 可维护，生效/失效时间正确", "商务/财务"),
    ("CFG-10", "基础配置", "阶段1-基础配置", "应收报价", "提拆派报价：复制历史报价", "已生效报价只读，复制后可编辑新单", "商务/财务"),
    ("CFG-11", "基础配置", "阶段1-基础配置", "应收报价", "库内操作费配置", "费用项关联正确", "财务"),
    ("CFG-12", "基础配置", "阶段1-基础配置", "服务配置", "工单规则/处理组/操作指令", "工单可自动分派到处理组", "客服主管"),
    # 提拆派与入库
    ("IN-01", "提拆派与入库", "阶段2-头程入库", "提拆派计划", "员工端创建提拆派计划", "计划创建成功，货件明细可维护", "调度"),
    ("IN-02", "提拆派与入库", "阶段2-头程入库", "提拆派计划", "计划导入/导出", "模板导入导出数据一致", "调度"),
    ("IN-03", "提拆派与入库", "阶段2-头程入库", "提拆派计划", "计划详情：轨迹/拆柜报告/附件/日志", "各 Tab 数据完整、日志有留痕", "调度"),
    ("IN-04", "提拆派与入库", "阶段2-头程入库", "提拆派计划", "计划状态推进至到仓", "状态变更字段校验通过", "调度"),
    ("IN-05", "提拆派与入库", "阶段2-头程入库", "提拆派计划", "拆柜完成触发报价匹配", "计费报告写入，命中则锁定并出账", "调度/财务"),
    ("IN-06", "提拆派与入库", "阶段2-头程入库", "提拆派计划", "组合仓计价路径", "多 FBA 仓拼柜按分区全包价分摊", "调度/财务"),
    ("IN-07", "提拆派与入库", "阶段2-头程入库", "提拆派计划", "散板计价路径", "未纳入组合仓价目表时改按散板", "调度/财务"),
    ("IN-08", "提拆派与入库", "阶段2-头程入库", "提拆派计划", "报价未命中场景", "不拦截拆柜、不锁定、不生成应收", "调度/财务"),
    ("IN-09", "提拆派与入库", "阶段2-头程入库", "货件管理", "拆柜后货件状态联动", "全部/未到仓/待出库视图一致", "调度"),
    ("IN-10", "提拆派与入库", "阶段2-头程入库", "货件管理", "问题件：放行/转留仓", "问题件处理流程正确", "调度/客服"),
    ("IN-11", "提拆派与入库", "阶段2-头程入库", "客户端", "客户端创建提拆派计划", "客户端与员工端数据一致", "客服/客户"),
    # 外仓收货
    ("OW-01", "外仓收货", "阶段3-外仓收货", "外仓收货", "提货回仓：待提货→提货中→已到仓", "司机/车牌/实收板数/收货 BOL 必填校验", "调度"),
    ("OW-02", "外仓收货", "阶段3-外仓收货", "外仓收货", "自送：待提货→已到仓", "货件状态：已预报→在库", "调度"),
    ("OW-03", "外仓收货", "阶段3-外仓收货", "外仓收货", "提货直送：→直送运输中→直送已到达", "货件 POD、出发/签收日期写入", "调度"),
    ("OW-04", "外仓收货", "阶段3-外仓收货", "外仓收货", "状态回退（跨级）", "回退原因必填，附件与字段按规则清理", "调度"),
    ("OW-05", "外仓收货", "阶段3-外仓收货", "外仓收货", "修改计划（各状态限制）", "不可改状态与字段限制符合 PRD", "调度"),
    ("OW-06", "外仓收货", "阶段3-外仓收货", "外仓收货", "取消计划", "计划与货件均为已取消终态", "调度"),
    ("OW-07", "外仓收货", "阶段3-外仓收货", "货件管理", "计划状态与货件状态联动", "货件动态展示货件状态非计划状态", "调度"),
    ("OW-08", "外仓收货", "阶段3-外仓收货", "客户端", "客户端创建外仓收货计划", "客户 BOL 上传，员工端可查看", "客服/客户"),
    # FBA
    ("FBA-01", "FBA出库派送", "阶段4-FBA出库", "FBA发货计划", "排货总览 / FBA仓汇总", "汇总数据与计划单一致", "调度"),
    ("FBA-02", "FBA出库派送", "阶段4-FBA出库", "FBA发货计划", "计划单创建与安排备货", "计划单与装车单正确关联", "调度"),
    ("FBA-03", "FBA出库派送", "阶段4-FBA出库", "FBA发货计划", "调度备货加货", "生成新 BOL 绑定原计划，不打加货标识", "调度"),
    ("FBA-04", "FBA出库派送", "阶段4-FBA出库", "FBA发货计划", "调度备货减货", "解除预备 BOL，不产生已作废", "调度"),
    ("FBA-05", "FBA出库派送", "阶段4-FBA出库", "装车单管理", "装车单：待装车→已装车→已发车", "装车单状态止于已发车", "调度/仓库"),
    ("FBA-06", "FBA出库派送", "阶段4-FBA出库", "FBA派送管理", "发车后 BOL 进入运输中", "计划/装车/BOL 分层状态一致", "调度"),
    ("FBA-07", "FBA出库派送", "阶段4-FBA出库", "FBA派送管理", "货件补录（已装车/运输中）", "按柜号+FBA Code 补录成功", "调度"),
    ("FBA-08", "FBA出库派送", "阶段4-FBA出库", "FBA派送管理", "单项退仓全流程", "通知退仓→执行→货件回在库", "调度/仓库"),
    ("FBA-09", "FBA出库派送", "阶段4-FBA出库", "FBA派送管理", "整车退仓全流程", "虚拟库位必填，装车单及 BOL 已装车", "调度/仓库"),
    ("FBA-10", "FBA出库派送", "阶段4-FBA出库", "FBA派送管理", "POD 上传与签收完结", "BOL 维度 POD 留存可查", "调度/客服"),
    # 其他出库
    ("OUT-01", "其他出库渠道", "阶段5-其他出库", "本地私仓卡派", "排货与出库流程", "私仓卡派单据状态正确推进", "调度"),
    ("OUT-02", "其他出库渠道", "阶段5-其他出库", "外州私仓卡派", "跨州私仓出库流程", "外州私仓出库可完成", "调度"),
    ("OUT-03", "其他出库渠道", "阶段5-其他出库", "自提单", "员工端自提管理与出库", "自提单创建至出库闭环", "调度/仓库"),
    ("OUT-04", "其他出库渠道", "阶段5-其他出库", "快递单", "快递出库流程", "快递单创建与状态跟踪", "调度"),
    ("OUT-05", "其他出库渠道", "阶段5-其他出库", "留仓货件", "留仓视图与转类型", "留仓货件可转类型并发起出库", "调度"),
    ("OUT-06", "其他出库渠道", "阶段5-其他出库", "留仓货件", "留仓货件发起出库指令", "出库指令下发后进入待出库", "调度"),
    ("OUT-07", "其他出库渠道", "阶段5-其他出库", "一件代发", "列表查询与导出", "数据查询导出正确", "运营"),
    # 财务
    ("FIN-01", "财务与报价", "阶段6-财务报价", "应收", "拆柜后自动出账", "报价匹配后应收明细按费用项拆分", "财务"),
    ("FIN-02", "财务与报价", "阶段6-财务报价", "应收", "组合仓费用分摊", "全包价减基础提拆价后按方数分摊正确", "财务"),
    ("FIN-03", "财务与报价", "阶段6-财务报价", "应收", "附加费（超重/超限/单询）", "边界场景金额与拦截符合规则", "财务"),
    ("FIN-04", "财务与报价", "阶段6-财务报价", "应收", "应收账单管理", "账单生成、明细勾稽正确", "财务"),
    ("FIN-05", "财务与报价", "阶段6-财务报价", "应收", "收款水单管理", "收款登记与核销", "财务"),
    ("FIN-06", "财务与报价", "阶段6-财务报价", "应收", "计划单费用汇总", "与 FBA 计划费用可核对", "财务"),
    ("FIN-07", "财务与报价", "阶段6-财务报价", "应付", "应付明细/汇总", "应付数据与业务单据一致", "财务"),
    ("FIN-08", "财务与报价", "阶段6-财务报价", "应付", "应付账单/付款凭证", "应付账单与付款闭环", "财务"),
    # 客户端
    ("CL-01", "客户端与服务", "阶段7-客户端", "客户端", "提拆派计划：创建/导入/详情", "客户端自助操作可用", "客服/客户"),
    ("CL-02", "客户端与服务", "阶段7-客户端", "客户端", "外仓收货计划：创建/BOL上传", "与员工端数据同步", "客服/客户"),
    ("CL-03", "客户端与服务", "阶段7-客户端", "客户端", "货件管理：全部/待出库/留仓/问题件", "各视图数据与员工端一致", "客服/客户"),
    ("CL-04", "客户端与服务", "阶段7-客户端", "客户端", "一件代发列表查询导出", "查询导出正确", "客服/客户"),
    ("CL-05", "客户端与服务", "阶段7-客户端", "工单", "客户端提工单→员工端处理→关闭", "工单全流程状态正确", "客服"),
    ("CL-06", "客户端与服务", "阶段7-客户端", "服务支持", "员工端工单管理/异常工单", "分派、回复、关闭留痕", "客服"),
    ("CL-07", "客户端与服务", "阶段7-客户端", "服务支持", "板标查询", "按板标号查货件信息", "客服/仓库"),
    ("CL-08", "客户端与服务", "阶段7-客户端", "账号", "客户端账号管理/个人中心", "子账号增删改、改密码", "客服/客户"),
    # PDA
    ("PDA-01", "PDA专项", "阶段2-头程入库", "PDA", "到仓入库：扫集装箱二维码", "到仓状态更新正确", "仓库"),
    ("PDA-02", "PDA专项", "阶段2-头程入库", "PDA", "到仓入库：扫库位+板标", "入库上架完成，货件在库", "仓库"),
    ("PDA-03", "PDA专项", "阶段2-头程入库", "PDA", "拆柜作业：开始/结束拆柜", "拆柜数量与异常录入正确", "仓库"),
    ("PDA-04", "PDA专项", "阶段2-头程入库", "PDA", "板标打印：Wi-Fi 打印机", "板标打印内容与格式正确", "仓库"),
    ("PDA-05", "PDA专项", "阶段2-头程入库", "PDA", "转库位：扫库位+板标", "转库后库位信息更新", "仓库"),
    ("PDA-06", "PDA专项", "阶段4-FBA出库", "PDA", "出库装车：扫板标确认", "BOL 不一致时系统报警", "仓库"),
    ("PDA-07", "PDA专项", "阶段4-FBA出库", "PDA", "POD反馈：拍照/升降机/签字", "POD 完整上传至 BOL", "仓库"),
    ("PDA-08", "PDA专项", "阶段4-FBA出库", "PDA", "FBA 备货加货/减货", "加货打标识、减货规则与员工端一致", "仓库"),
    ("PDA-09", "PDA专项", "阶段4-FBA出库", "PDA", "按指令执行退仓", "退仓状态回写正确", "仓库"),
    ("PDA-10", "PDA专项", "阶段5-其他出库", "PDA", "自提核验：扫自提码出库", "身份核验通过后出库完成", "仓库"),
    ("PDA-11", "PDA专项", "阶段2-头程入库", "PDA", "主管督导页面", "督导视图与异常处理可用", "仓库主管"),
    # 端到端
    ("E2E-A1", "端到端剧本", "阶段8-端到端", "剧本A", "客户下单提拆派计划", "客户端/员工端计划创建成功", "客服/调度"),
    ("E2E-A2", "端到端剧本", "阶段8-端到端", "剧本A", "到仓→PDA入库拆柜", "仓库作业完成，状态联动", "仓库/调度"),
    ("E2E-A3", "端到端剧本", "阶段8-端到端", "剧本A", "拆柜报价出账", "应收明细生成，财务可核对", "财务"),
    ("E2E-A4", "端到端剧本", "阶段8-端到端", "剧本A", "创建FBA计划→备货装车", "计划/装车/BOL 关联正确", "调度/仓库"),
    ("E2E-A5", "端到端剧本", "阶段8-端到端", "剧本A", "PDA发车→POD签收", "派送完结，POD 可查", "仓库/调度"),
    ("E2E-A6", "端到端剧本", "阶段8-端到端", "剧本A", "客户端查看轨迹", "客户端状态与员工端一致", "客服/客户"),
    ("E2E-B1", "端到端剧本", "阶段8-端到端", "剧本B", "外仓收货（提货直送）", "直送全流程完成", "调度"),
    ("E2E-B2", "端到端剧本", "阶段8-端到端", "剧本B", "问题件拦截→工单处理", "异常处置闭环", "调度/客服"),
    ("E2E-B3", "端到端剧本", "阶段8-端到端", "剧本B", "FBA单项退仓→回在库", "退仓后货件可重新排货", "调度/仓库"),
    ("E2E-B4", "端到端剧本", "阶段8-端到端", "剧本B", "重新排货出库", "同一货件二次出库成功", "调度/仓库"),
    ("E2E-C1", "端到端剧本", "阶段8-端到端", "非功能", "多 PDA 同时扫码", "无明显卡顿或数据冲突", "仓库/IT"),
    ("E2E-C2", "端到端剧本", "阶段8-端到端", "非功能", "三端同一货件状态一致性", "Admin/Client/PDA 状态一致", "全员"),
]

MILESTONES = [
    ("阶段0", "6/15-6/16", "启动会完成，账号可用", "2026-06-16"),
    ("阶段1", "6/17-6/20", "基础配置验收签字", "2026-06-20"),
    ("阶段2", "6/21-6/25", "头程入库链路双签", "2026-06-25"),
    ("阶段3", "6/26-6/29", "外仓收货三方式验收", "2026-06-29"),
    ("阶段4", "6/30-7/5", "FBA 主流程+退仓验收", "2026-07-05"),
    ("阶段5", "7/6-7/8", "各出库渠道各1单跑通", "2026-07-08"),
    ("阶段6", "7/1-7/7", "财务费用样例签字", "2026-07-07"),
    ("阶段7", "7/6-7/9", "试点客户自助+工单", "2026-07-09"),
    ("阶段8", "7/10-7/12", "端到端回归报告", "2026-07-12"),
    ("阶段9", "7/13-7/15", "培训完成，正式上线", "2026-07-15"),
]

SIGNOFF_ITEMS = [
    ("阶段1-基础配置", "主数据与系统配置验收通过", "2026-06-20"),
    ("阶段2-头程入库", "提拆派+PDA入库拆柜闭环", "2026-06-25"),
    ("阶段3-外仓收货", "三种提货方式验收通过", "2026-06-29"),
    ("阶段4-FBA出库", "FBA计划派送+退仓验收", "2026-07-05"),
    ("阶段5-其他出库", "私仓/自提/快递/留仓验收", "2026-07-08"),
    ("阶段6-财务", "费用出账与对账样例确认", "2026-07-07"),
    ("阶段7-客户端", "试点客户验收通过", "2026-07-09"),
    ("阶段8-端到端", "回归测试 Go 评估", "2026-07-12"),
    ("阶段9-上线", "正式上线 Go-Live", "2026-07-15"),
]

COL_COUNT = 11  # max columns used across sections


def style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def write_section_title(ws, row, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COL_COUNT)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center")
    for c in range(1, COL_COUNT + 1):
        ws.cell(row=row, column=c).border = BORDER


def apply_borders(ws, start_row, end_row, col_count):
    for r in range(start_row, end_row + 1):
        for c in range(1, col_count + 1):
            ws.cell(row=r, column=c).border = BORDER


def add_result_validation(ws, col_letter, start_row, end_row):
    dv = DataValidation(type="list", formula1=RESULT_OPTIONS, allow_blank=True)
    dv.error = "请从下拉列表选择验收结果"
    dv.errorTitle = "无效输入"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def add_list_validation(ws, col_letter, start_row, end_row, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def add_conditional_result_colors(ws, col_letter, start_row, end_row):
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    for text, color in [("通过", "C6EFCE"), ("不通过", "FFC7CE"), ("阻塞", "FFEB9C"), ("待验证", "DDEBF7")]:
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'${col_letter}${start_row}="{text}"'],
                fill=PatternFill("solid", fgColor=color),
            ),
        )


def write_title_block(ws, col_count, checklist_hint):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws.cell(row=1, column=1, value="MEEKOO-TMS 业务验收检查表").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws.cell(row=2, column=1, value=(
        "验收周期：2026年6月15日 — 2026年7月15日 ｜ 目标上线：2026年7月15日 ｜ "
        "系统范围：员工端 + 客户端 + PDA ｜ 验收结果：通过/不通过/阻塞/待验证/N/A"
    )).font = META_FONT
    ws.cell(row=2, column=1).alignment = WRAP

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=col_count)
    ws.cell(row=3, column=1, value=checklist_hint).font = META_FONT
    ws.cell(row=3, column=1).alignment = WRAP
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 36


def build_overview_sheet(wb):
    ws = wb.active
    ws.title = "总览"

    widths = [10, 12, 28, 14, 14, 10, 12, 12, 20, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    write_title_block(
        ws, COL_COUNT,
        "填写说明：① 验收进度与阶段签字在本表维护；② 具体检查项在「清单」Sheet 填写；"
        "③ 不通过/阻塞项在下方缺陷跟踪登记，清单备注栏填写 BUG 编号；"
        "④ P0=主链路阻断，P1=有临时方案，P2=可上线后迭代。",
    )

    row = 5

    write_section_title(ws, row, "一、验收进度总览")
    row += 1
    ms_headers = ["阶段", "时间范围", "里程碑", "计划完成日", "实际完成日", "验收结果", "业务签字", "IT/产品签字", "备注"]
    for c, h in enumerate(ms_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(ms_headers))
    row += 1
    ms_start = row
    for m in MILESTONES:
        ws.append([*m, "", "", "", "", ""])
        row += 1
    apply_borders(ws, ms_start - 1, row - 1, len(ms_headers))
    add_result_validation(ws, "F", ms_start, row - 1)

    row += 1

    write_section_title(ws, row, "二、缺陷跟踪")
    row += 1
    bug_headers = [
        "编号", "发现日期", "场景编号", "模块", "问题描述", "级别", "状态",
        "发现人", "修复人", "修复日期", "复验结果",
    ]
    for c, h in enumerate(bug_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(bug_headers))
    row += 1
    bug_start = row
    for i in range(1, 31):
        ws.append([f"BUG-{i:03d}", "", "", "", "", "", "待修复", "", "", "", ""])
        row += 1
    apply_borders(ws, bug_start - 1, row - 1, len(bug_headers))
    add_list_validation(ws, "F", bug_start, row - 1, LEVEL_OPTIONS)
    add_list_validation(ws, "G", bug_start, row - 1, STATUS_OPTIONS)
    add_result_validation(ws, "K", bug_start, row - 1)

    row += 1

    write_section_title(ws, row, "三、阶段签字确认")
    row += 1
    sign_headers = [
        "阶段", "里程碑描述", "计划日期", "实际日期", "P0数", "P1数",
        "业务负责人", "业务签字", "签字日期", "产品/IT", "结论(Go/No-Go)",
    ]
    for c, h in enumerate(sign_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(sign_headers))
    row += 1
    sign_start = row
    for item in SIGNOFF_ITEMS:
        ws.append([*item, "", 0, 0, "", "", "", "", ""])
        row += 1
    apply_borders(ws, sign_start - 1, row - 1, len(sign_headers))

    ws.freeze_panes = "A5"
    return ws


def build_checklist_sheet(wb):
    ws = wb.create_sheet("清单")

    widths = [10, 12, 14, 14, 34, 30, 10, 10, 12, 10, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    write_title_block(
        ws, len(CHECK_HEADERS),
        "填写说明：① 按「验收分类」或「验收阶段」筛选本部门负责项；"
        "② 不通过/阻塞项在「总览」Sheet 登记缺陷，本表备注栏填写 BUG 编号；"
        "③ 每阶段完成后在「总览」Sheet 签字确认。",
    )

    row = 5
    write_section_title(ws, row, "验收检查清单（共 {} 项）".format(len(ALL_CHECK_ITEMS)))
    row += 1
    for c, h in enumerate(CHECK_HEADERS, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(CHECK_HEADERS))
    row += 1
    check_start = row
    for item in ALL_CHECK_ITEMS:
        ws.append([*item, "", "", "待验证", ""])
        row += 1
    apply_borders(ws, check_start - 1, row - 1, len(CHECK_HEADERS))
    for r in range(check_start, row):
        for c in (5, 6, 11):
            ws.cell(row=r, column=c).alignment = WRAP
    add_result_validation(ws, "J", check_start, row - 1)
    add_conditional_result_colors(ws, "J", check_start, row - 1)
    ws.auto_filter.ref = f"A{check_start - 1}:K{row - 1}"
    ws.freeze_panes = f"A{check_start}"
    return ws


def main():
    wb = Workbook()
    build_overview_sheet(wb)
    build_checklist_sheet(wb)
    target = OUT
    try:
        wb.save(target)
    except PermissionError:
        target = OUT.with_name(OUT.stem + "-最新" + OUT.suffix)
        wb.save(target)
        print(f"原文件被占用，已改写到: {target}")
    print(f"已生成: {target}（总览 + 清单，共 {len(ALL_CHECK_ITEMS)} 条验收项）")


if __name__ == "__main__":
    main()
