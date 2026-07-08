# -*- coding: utf-8 -*-
"""
生成「提拆派报价」Excel 导入模板（多 Sheet，含附加费）。

用法：
  python admin/docs/_generate_pricing_import_template.py

输出：
  admin/templates/提拆派报价导入模板.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = Path(__file__).resolve().parent.parent / "templates"
OUT = OUT_DIR / "提拆派报价导入模板.xlsx"

CONTAINER_TYPES_DATA = [
    ("20GP", "20英尺普柜"),
    ("40GP", "40英尺普柜"),
    ("40HQ", "40英尺高柜"),
    ("45HC", "45英尺高柜"),
    ("45HQ", "45英尺高柜"),
]
CONTAINER_TYPES = [code for code, _ in CONTAINER_TYPES_DATA]

# 与 pricing-pickup-delivery-edit.html · PRICING_FEE_ITEMS 对齐（仅名称+环节，不含编码）
FEE_ITEMS = [
    ("提柜费", "提柜"),
    ("拆柜提拆费", "拆柜"),
    ("卡派费", "卡派"),
    ("整柜直送运费", "卡派"),
    ("整柜快递运费", "卡派"),
    ("组合仓派送费", "卡派"),
    ("散板派送费", "卡派"),
    ("其他费用", "其他"),
]
FEE_ITEM_NAMES = [name for name, _ in FEE_ITEMS]

DS_SHEET = "数据源"

FEE_BINDINGS = [
    ("整柜直送", "整柜直送运费", "对应 02-整柜直送 价目"),
    ("整柜快递", "整柜快递运费", "对应 03-整柜快递·散板自提·散板快递 中整柜快递"),
    ("组合仓派送", "组合仓派送费", "对应 05-组合仓派送"),
    ("散板派送", "散板派送费", "对应 FBA/Walmart/私卡派/自提/快递散板价目"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
HINT_FONT = Font(size=10, color="666666")
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def write_rows(ws, start_row, rows):
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = WRAP
            cell.border = BORDER


def autosize_columns(ws, min_width=10, max_width=48):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v).split("\n")[0]))
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_width), max_width)


def add_list_validation(ws, cell_range, source_formula):
    dv = DataValidation(type="list", formula1=source_formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def sheet_data_source(wb):
    ws = wb.create_sheet(DS_SHEET)
    row = 1
    title = ws.cell(row=row, column=1, value="费用项（来自基础数据 · 费用项配置）")
    title.font = SECTION_FONT
    row += 1
    for col, title_text in enumerate(["费用项名称", "环节"], start=1):
        ws.cell(row=row, column=col, value=title_text)
    style_header_row(ws, row, 2)
    fee_start = row + 1
    for i, (name, stage) in enumerate(FEE_ITEMS):
        r = fee_start + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=stage)
        for c in range(1, 3):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER
    fee_end = fee_start + len(FEE_ITEMS) - 1

    row = fee_end + 2
    ws.cell(row=row, column=1, value="柜型").font = SECTION_FONT
    row += 1
    for col, title_text in enumerate(["柜型", "说明"], start=1):
        ws.cell(row=row, column=col, value=title_text)
    style_header_row(ws, row, 2)
    ct_start = row + 1
    for i, (code, desc) in enumerate(CONTAINER_TYPES_DATA):
        r = ct_start + i
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=desc)
        for c in range(1, 3):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER
    ct_end = ct_start + len(CONTAINER_TYPES_DATA) - 1

    hint_row = ct_end + 2
    ws.cell(row=hint_row, column=1, value="说明：本 Sheet 为下拉数据源，请勿删除；正式导入时以系统「费用项配置」为准。").font = HINT_FONT
    autosize_columns(ws)
    return {
        "fee_names": f"={DS_SHEET}!$A${fee_start}:$A${fee_end}",
        "container_types": f"={DS_SHEET}!$A${ct_start}:$A${ct_end}",
    }


def sheet_instructions(wb):
    ws = wb.active
    ws.title = "填写说明"
    ws["A1"] = "提拆派报价 · Excel 导入模板"
    ws["A1"].font = TITLE_FONT
    lines = [
        "",
        "【使用说明】",
        "1. 请按 Sheet 顺序填写；带 * 的字段为必填。",
        "2. 仓库代码多个请用 / 分隔，如 ONT8/LGB8/LAX9。",
        "3. 柜型、费用项名称请参照「数据源」Sheet；价目表柜型列仅填写需要报价的柜型，未使用的可留空或填 0。",
        "4. 整柜直送、组合仓派送：全包价（含提拆+派送）；「基础提拆价格」在 04-拆柜-提拆费 填写，同一拆柜计划整柜仅计一次。",
        "5. 散板派送：FBA卡派、Walmart卡派 按分区+仓库代码维护板单价；私卡派 按距离+板数阶梯维护。",
        "   整柜快递、散板自提、散板快递 合并在同一 Sheet，通过「分类」区分。",
        "6. 附加费：",
        "   · 超出按量加收：整柜箱数/拼仓目的地数 超过阈值后，按超出量 × 单价计费。",
        "   · 触发后固定加收：命中条件后一次性加收固定金额（可按柜型限定）。",
        "   · 重量阶梯：同一费用名称可填多行，按柜型+预报总柜重下限匹配最高一档。",
        "7. 时间格式：YYYY-MM-DD HH:MM:SS，如 2026-01-01 00:00:00。",
        "8. 结算币种：USD / CNY / EUR / GBP。",
        "9. 关联费用项：01-基础信息 绑定各价目 Tab 的费用项；04-拆柜-提拆费 单独绑定提拆费项；选项见「数据源」。",
        "",
        "【Sheet 清单】",
        "数据源 · 费用项名称/环节、柜型/说明（位于最后，下拉引用，请勿删除）",
        "01-基础信息 · 报价名称、适用客户、生效/失效时间、结算币种、关联费用项",
        "02-整柜直送 · 分区 + 仓库代码 + 各柜型全包价",
        "03-整柜快递·散板自提·散板快递 · 整柜快递（柜型+单价）/ 散板自提 / 散板快递（板单价）",
        "04-拆柜-提拆费 · 基础提拆价格 + 关联费用项（与编辑页拆柜 Tab 一致）",
        "05-组合仓派送 · 分区 + 仓库代码 + 各柜型全包价",
        "06-散板-FBA卡派 · 分区 + 仓库代码 + 板单价",
        "07-散板-Walmart卡派 · 分区 + 仓库代码 + 板单价",
        "08-散板-私卡派 · 距离范围 + 板数范围 + 报价方式 + 价格",
        "09-附加费 · 附加费规则（含重量阶梯明细）",
    ]
    for i, text in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=text).font = HINT_FONT if i > 2 else Font(size=11)
    ws.column_dimensions["A"].width = 92


def sheet_basic(wb, ds):
    ws = wb.create_sheet("01-基础信息")
    headers = ["字段*", "填写值*", "说明"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    rows = [
        ["报价名称", "2026标准价目表", "必填，报价单显示名称"],
        ["适用客户", "所有客户", "所有客户 或 指定客户"],
        ["指定客户", "", "适用客户=指定客户时必填，填写客户代码（非客户名称），多个用英文逗号分隔，如：tx,dm"],
        ["生效时间", "2026-01-01 00:00:00", "必填，格式 YYYY-MM-DD HH:MM:SS"],
        ["失效时间", "2026-12-31 23:59:59", "必填，须晚于生效时间"],
        ["结算币种", "USD", "USD / CNY / EUR / GBP"],
    ]
    write_rows(ws, 2, rows)

    fee_section_row = len(rows) + 3
    ws.cell(row=fee_section_row, column=1, value="关联费用项（出账时写入应收）")
    ws.cell(row=fee_section_row, column=1).font = SECTION_FONT
    ws.merge_cells(start_row=fee_section_row, start_column=1, end_row=fee_section_row, end_column=3)

    fee_header_row = fee_section_row + 1
    fee_headers = ["价目分类*", "关联费用项*", "说明"]
    for col, title in enumerate(fee_headers, start=1):
        ws.cell(row=fee_header_row, column=col, value=title)
    style_header_row(ws, fee_header_row, len(fee_headers))

    fee_rows = [[cat, name, hint] for cat, name, hint in FEE_BINDINGS]
    write_rows(ws, fee_header_row + 1, fee_rows)

    dv_scope = DataValidation(type="list", formula1='"所有客户,指定客户"', allow_blank=True)
    ws.add_data_validation(dv_scope)
    dv_scope.add(ws["B3"])
    add_list_validation(ws, ws["B7"], '"USD,CNY,EUR,GBP"')
    first_fee_data_row = fee_header_row + 1
    last_fee_data_row = first_fee_data_row + len(FEE_BINDINGS) - 1
    add_list_validation(ws, f"B{first_fee_data_row}:B{last_fee_data_row}", ds["fee_names"])
    autosize_columns(ws)


def sheet_unpack_base(wb, ds):
    ws = wb.create_sheet("04-拆柜-提拆费")
    headers = ["基础提拆价格*", "关联费用项*", "说明"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    write_rows(ws, 2, [[
        1400,
        "拆柜提拆费",
        "同一拆柜计划内整柜仅计一次；与组合仓/散板派送无关",
    ]])
    add_list_validation(ws, "B2", ds["fee_names"])
    autosize_columns(ws)


def sheet_direct(wb, ds):
    ws = wb.create_sheet("02-整柜直送")
    headers = ["分区*", "仓库代码*", *CONTAINER_TYPES, "备注"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    samples = [
        ["南加州 (一档)", "ONT8/LGB8/LAX9/SBD1/POC1/POC3", "", "", 1300, 1400, "", "含车架费4天，含1次车架分离费"],
        ["南加州 (二档)", "SBD1/SBD2/LGB6/LGB9/ONT9", "", "", 1350, 1450, "", "地板货或全托盘货，不得混装"],
        ["北加州 (一档)", "SMF3/SCK4/SCK1/SJC7/MCE1/OAK3", "", "", 2500, 2600, "", "长途直送线路"],
    ]
    write_rows(ws, 2, samples)
    autosize_columns(ws)


def sheet_express_and_loose_flat(wb, ds):
    ws = wb.create_sheet("03-整柜快递·散板自提·散板快递")
    headers = ["分类*", "柜型", "价格*", "备注"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    write_rows(ws, 2, [
        ["整柜快递", "40HQ", 11, ""],
        ["整柜快递", "45HQ", 0, ""],
        ["散板-自提", "", 0, "自提货件按板计费"],
        ["散板-快递", "", 0, "快递货件按板计费"],
    ])
    add_list_validation(ws, "A2:A200", '"整柜快递,散板-自提,散板-快递"')
    add_list_validation(ws, "B2:B200", ds["container_types"])
    autosize_columns(ws)


def sheet_combo(wb, ds):
    ws = wb.create_sheet("05-组合仓派送")
    headers = ["分区*", "仓库代码*", *CONTAINER_TYPES, "备注"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    samples = [
        ["一区", "ONT8/LGB8/SBD1/LAX9/POC1/POC2/POC3/LAX2/LAX1", "", "", 1850, 2050, "", "拆柜后1-5自然日入仓"],
        ["二区", "IUSJ/IUSQ/IUSP/IUSW/IUTI/XLX7/ONT9/LGB4/LGB6/SBD2/SNA4", "", "", 2200, 2400, "", "拆柜后1-5自然日入仓"],
        ["三区", "SCK1/SCK4/SMF3/GYR2/GYR3/LAS1/VGT2/MIT2/GEU2/GEU3/GEU5/MCE1/FAT2/LAS6/OAK3/SJC7/QXY8/IUTE/SMF1", "", "", 2950, 3050, "", "拆柜后3-7自然日入仓"],
    ]
    write_rows(ws, 2, samples)
    autosize_columns(ws)


def sheet_loose_amazon(wb):
    ws = wb.create_sheet("06-散板-FBA卡派")
    headers = ["分区*", "仓库代码*", "板单价*", "备注"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    samples = [
        ["CA", "ONT8 / LGB8 / LAX9 / SBD1", 25, ""],
        ["CA", "KRB1 / KRB4 / ONT9 / LGB6 / LGB9 / SBD2 / KRB7", 35, ""],
        ["NV", "LAS1 / VGT2", 55, ""],
    ]
    write_rows(ws, 2, samples)
    autosize_columns(ws)


def sheet_loose_walmart(wb):
    ws = wb.create_sheet("07-散板-Walmart卡派")
    headers = ["分区*", "仓库代码*", "板单价*", "备注"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    write_rows(ws, 2, [
        ["CA", "LAX1 / LAX2", 45, "出库后1-3个工作日签收"],
        ["CA/AZ", "SMF1 / PHX1", 85, "出库后3-5个工作日签收"],
        ["KS", "KS1", 245, "出库后5-8个工作日签收"],
    ])
    autosize_columns(ws)


def sheet_loose_private(wb):
    ws = wb.create_sheet("08-散板-私卡派")
    headers = [
        "距离下限(Mile)*", "距离上限(Mile)*",
        "板数下限*", "板数上限*",
        "报价方式*", "一口价", "首托价", "续托价", "单托价", "说明",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    rows = [
        [0, 30, 1, 12, "首托+续托阶梯", "", 150, 10, "", "1-12板：首托150，续托10/板"],
        [0, 30, 13, 26, "一口价", 450, "", "", "", "13-26板整车一口价"],
        [31, 50, 1, 12, "首托+续托阶梯", "", 200, 10, "", ""],
        [31, 50, 13, 26, "一口价", 700, "", "", "", ""],
    ]
    write_rows(ws, 2, rows)
    dv = DataValidation(
        type="list",
        formula1='"一口价,首托+续托阶梯,按固定单托计费"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add("E2:E500")
    autosize_columns(ws)


def sheet_surcharge(wb, ds):
    ws = wb.create_sheet("09-附加费")
    headers = [
        "费用名称*", "规则类型*",
        "计量指标", "阈值", "超出单价/固定金额",
        "适用柜型", "重量下限(LBS)", "阶梯加收金额",
        "备注",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    rows = [
        ["超箱费", "超出按量加收", "整柜箱数", 1200, 0.2, "", "", "", "超过1200箱，超出部分0.2/箱"],
        ["超地址费", "超出按量加收", "拼仓派送目的地数(直送柜不收取)", 10, 50, "", "", "", "自提/私卡派/快递各算1个目的地；亚马逊/沃尔玛每仓算1个"],
        ["固定加收", "触发后固定加收", "柜型", "", 80, "40HQ / 45HQ", "", "", "命中柜型后固定加收"],
        ["提柜超重", "重量阶梯", "预报总柜重", "", "", "20GP / 20HQ", 38000, 200, "17.2吨"],
        ["提柜超重", "重量阶梯", "预报总柜重", "", "", "40GP / 40HQ", 43000, 150, "19.5吨"],
        ["提柜超重", "重量阶梯", "预报总柜重", "", "", "40HQ / 45HQ", 48000, 250, "21.8吨"],
        ["提柜超重", "重量阶梯", "预报总柜重", "", "", "40HQ / 45HQ", 52000, 500, "23.58吨"],
        ["超板费", "超出按量加收", "整柜箱数", 0, 0, "", "", "", "可按需填写"],
        ["燃油附加费", "触发后固定加收", "柜型", "", 0, "", "", "", "可按需填写"],
    ]
    write_rows(ws, 2, rows)
    add_list_validation(ws, "B2:B500", '"超出按量加收,触发后固定加收,重量阶梯"')
    add_list_validation(ws, "C2:C500", '"整柜箱数,拼仓派送目的地数(直送柜不收取),预报总柜重,柜型"')
    add_list_validation(ws, "A2:A500", '"超箱费,超地址费,提柜超重,固定加收,超板费,超方费,操作费,燃油附加费,分拣费,其他附加费"')
    add_list_validation(ws, "F2:F500", ds["container_types"])
    autosize_columns(ws, max_width=36)
    ws.freeze_panes = "A2"


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheet_instructions(wb)
    ds = sheet_data_source(wb)
    sheet_basic(wb, ds)
    sheet_direct(wb, ds)
    sheet_express_and_loose_flat(wb, ds)
    sheet_unpack_base(wb, ds)
    sheet_combo(wb, ds)
    sheet_loose_amazon(wb)
    sheet_loose_walmart(wb)
    sheet_loose_private(wb)
    sheet_surcharge(wb, ds)
    # 数据源需先创建供下拉引用，生成完成后移到最后（附加费之后）
    ds_index = wb.sheetnames.index(DS_SHEET)
    wb.move_sheet(DS_SHEET, offset=len(wb.sheetnames) - ds_index - 1)
    wb.save(OUT)
    print(f"已生成: {OUT}")


if __name__ == "__main__":
    main()
